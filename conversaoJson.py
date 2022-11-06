import openpyxl as op
import json
import os
import datetime as dt

# Carrega o arquivo de planilha
planilha = op.load_workbook('cobranca.xlsx')
aba = planilha.active
t = aba.max_row
print('linhas ', t)
a = 2
clientes = {}
while a < t:
    print('linha ', a, '\n')
    nome = aba.cell(row=a, column=1).value
    vencimento = aba.cell(row=a, column=2).value
    vencimento = vencimento.strftime('%d/%m')
    telefone = aba.cell(row=a, column=3).value
    cdbarras = aba.cell(row=a, column=4).value
    y = {
        'nome': nome,
        'vencimento': vencimento,
        'telefone': telefone,
        'codbarras': cdbarras
    }
    clientes[a-2] = y
    print(y)
    a += 1

if os.path.exists('cobranca.json'):
    os.remove('cobranca.json')
clientes = json.dumps(clientes, indent=4)
with open('cobranca.json', 'w') as f:
    f.write(clientes)
